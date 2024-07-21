import os
from openpyxl import Workbook, load_workbook

base_dir = os.path.dirname(os.path.abspath(__file__))
assets_dir = os.path.join(base_dir, 'assets')
file_path = os.path.join(assets_dir, 'planilha.xlsx')

wb = Workbook()

ws = wb.active
ws.title = 'Planilha'

ws['A1'] = 'Nome'
ws['B1'] = 'Idade'
ws['C1'] = 'Sexo'

ws['A2'] = 'João'
ws['B2'] = 25
ws['C2'] = 'M'

ws['A3'] = 'Maria'
ws['B3'] = 30
ws['C3'] = 'F'

wb.save(file_path)

wb = load_workbook(file_path)

ws = wb.active

class Pessoa:
    def __init__(self, nome, idade, sexo):
        self.nome = nome
        self.idade = idade
        self.sexo = sexo

pessoas = []

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3, values_only=True):
    nome, idade, sexo = row
    pessoa = Pessoa(nome, idade, sexo)
    pessoas.append(pessoa)

for pessoa in pessoas:
    print(f"Nome: {pessoa.nome}, Idade: {pessoa.idade}, Sexo: {pessoa.sexo}")
